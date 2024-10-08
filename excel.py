import openpyxl
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import os
import sys

# Function to load a workbook when the user selects an Excel file
def load_excel_file():
    global wb, file_path, sheet_names, sheet_var
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    
    if file_path:
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet_names = wb.sheetnames
            sheet_var.set(sheet_names[0])  # Set to the first sheet by default
            sheet_dropdown['values'] = sheet_names  # Update the dropdown with the new sheet names
            update_sheet()
            messagebox.showinfo("File Loaded", f"Excel file '{os.path.basename(file_path)}' loaded successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file: {str(e)}")

# Function to update the sheet variable
def update_sheet():
    global sheet
    selected_sheet = sheet_var.get()
    if wb:
        sheet = wb[selected_sheet]

# Create Tkinter GUI Setup
root = tk.Tk()
root.title("School Details Management")

# Add color schemes
bg_color = "#FFFFFF"  # White background
label_color = "#000000"  # Black for labels
entry_bg = "#F5F5F5"  # Light gray for entries
button_bg = "#1E90FF"  # DodgerBlue for buttons
button_fg = "white"  # White text for buttons
root.configure(bg=bg_color)

# Create main frame
main_frame = tk.Frame(root, bg=bg_color)
main_frame.pack(fill=tk.BOTH, expand=1, padx=20, pady=20)

# Tkinter Variables
sr_no_var = tk.StringVar()
city_var = tk.StringVar()
school_name_var = tk.StringVar()
principal_name_var = tk.StringVar()
address_var = tk.StringVar()
phone_number_var = tk.StringVar()
email_var = tk.StringVar()

# Button to select an Excel file
tk.Button(main_frame, text="Select Excel File", command=load_excel_file, bg=button_bg, fg=button_fg, font=('Arial', 10)).grid(row=0, column=0, padx=10, pady=5)

# Create a dropdown for sheet selection
sheet_var = tk.StringVar()  # Variable to hold selected sheet name
sheet_dropdown = ttk.Combobox(main_frame, textvariable=sheet_var, font=('Arial', 12), state="readonly")
sheet_dropdown.grid(row=1, column=1, padx=10, pady=5, sticky="w")
tk.Label(main_frame, text="Select Sheet:", bg=bg_color, fg=label_color, font=('Arial', 12, 'bold')).grid(row=1, column=0, padx=10, pady=5, sticky="e")
sheet_dropdown.bind("<<ComboboxSelected>>", lambda e: update_sheet())

# Column Indexes
SR_NO_COL = 1           # Sr.No column
CITY_COL = 2            # City column
SCHOOL_NAME_COL = 3     # School Name column
PRINCIPAL_NAME_COL = 4  # Principal Name column
ADDRESS_COL = 5         # Address column
PHONE_NUMBER_COL = 6    # Phone Number column
EMAIL_COL = 7           # Email column

# Function to fetch school details by Sr.No
def fetch_school_details():
    if not file_path:
        messagebox.showwarning("File Error", "Please select an Excel file first.")
        return
    sr_no = sr_no_var.get()
    if not sr_no:
        messagebox.showwarning("Input Error", "Sr.No must be provided to fetch details.")
        return

    # Reset field colors
    set_field_colors("")

    try:
        sr_no = int(sr_no)
    except ValueError:
        messagebox.showwarning("Input Error", "Sr.No must be an integer.")
        return

    for row in sheet.iter_rows(min_row=2):
        current_sr_no = row[SR_NO_COL - 1].value

        # Skip cells with formulas
        if isinstance(current_sr_no, str) and current_sr_no.startswith("="):
            continue

        if current_sr_no is not None:
            try:
                if int(str(current_sr_no).strip()) == sr_no:
                    city_var.set(row[CITY_COL - 1].value)
                    school_name_var.set(row[SCHOOL_NAME_COL - 1].value)
                    principal_name_var.set(row[PRINCIPAL_NAME_COL - 1].value)
                    address_var.set(row[ADDRESS_COL - 1].value)
                    phone_number_var.set(row[PHONE_NUMBER_COL - 1].value)
                    email_var.set(row[EMAIL_COL - 1].value)
                    set_field_colors("success")  # Indicate fetch success
                    return
            except ValueError:
                continue

    messagebox.showwarning("Fetch Error", f"No entry found with Sr.No: {sr_no}")
    set_field_colors("error")

# Function to update or insert school details
def update_school():
    sr_no = sr_no_var.get()
    city = city_var.get()
    school_name = school_name_var.get()
    principal_name = principal_name_var.get()
    address = address_var.get()
    phone_number = phone_number_var.get()
    email = email_var.get()

    if not sr_no or not city or not school_name or not principal_name or not address:
        messagebox.showwarning("Input Error", "Sr.No, City, School Name, Principal Name, and Address fields must be filled.")
        set_field_colors("error")
        return

    try:
        sr_no = int(sr_no)
    except ValueError:
        messagebox.showwarning("Input Error", "Sr.No must be an integer.")
        set_field_colors("error")
        return

    school_found = False

    # Update school details if Sr.No already exists
    for row in sheet.iter_rows(min_row=2):
        if str(row[SR_NO_COL - 1].value).strip() == str(sr_no):
            row[CITY_COL - 1].value = city
            row[SCHOOL_NAME_COL - 1].value = school_name
            row[PRINCIPAL_NAME_COL - 1].value = principal_name
            row[ADDRESS_COL - 1].value = address
            row[PHONE_NUMBER_COL - 1].value = phone_number
            row[EMAIL_COL - 1].value = email
            school_found = True
            break

    # Insert new row if Sr.No does not exist
    if not school_found:
        new_row = [sr_no, city, school_name, principal_name, address, phone_number, email]
        sheet.append(new_row)

    # Save the workbook and show success message
    wb.save(file_path)
    messagebox.showinfo("Success", "School details have been successfully saved.")
    set_field_colors("success")

# Function to delete school details
def delete_school():
    sr_no = sr_no_var.get()
    if not sr_no:
        messagebox.showwarning("Input Error", "Sr.No must be provided to delete details.")
        set_field_colors("error")
        return

    try:
        sr_no = int(sr_no)
    except ValueError:
        messagebox.showwarning("Input Error", "Sr.No must be an integer.")
        set_field_colors("error")
        return

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        current_sr_no = row[SR_NO_COL - 1].value

        # Skip cells with formulas
        if isinstance(current_sr_no, str) and current_sr_no.startswith("="):
            continue

        if current_sr_no is not None and int(str(current_sr_no).strip()) == sr_no:
            sheet.delete_rows(row_idx)
            wb.save(file_path)
            messagebox.showinfo("Delete Success", f"Deleted details for Sr.No: {sr_no}")
            set_field_colors("success")
            return

    messagebox.showwarning("Delete Error", f"No entry found with Sr.No: {sr_no}")
    set_field_colors("error")

# Function to set field colors based on success or error
def set_field_colors(status):
    if status == "success":
        color = "#90EE90"  # Light green for success
    elif status == "error":
        color = "#FFCCCB"  # Light red for error
    else:
        color = entry_bg  # Reset to original entry background color

    # Set the color for all entry fields
    sr_no_entry.configure(bg=color)
    city_entry.configure(bg=color)
    school_name_entry.configure(bg=color)
    principal_name_entry.configure(bg=color)
    address_entry.configure(bg=color)
    phone_number_entry.configure(bg=color)
    email_entry.configure(bg=color)

# Create input fields and labels
tk.Label(main_frame, text="Sr.No:", bg=bg_color, fg=label_color, font=('Arial', 12)).grid(row=2, column=0, padx=10, pady=5, sticky="e")
sr_no_entry = tk.Entry(main_frame, textvariable=sr_no_var, bg=entry_bg, font=('Arial', 12))
sr_no_entry.grid(row=2, column=1, padx=10, pady=5, sticky="w")
tk.Button(main_frame, text="Fetch Details", command=fetch_school_details, bg=button_bg, fg=button_fg, font=('Arial', 10)).grid(row=2, column=2, padx=10, pady=5)

tk.Label(main_frame, text="City:", bg=bg_color, fg=label_color, font=('Arial', 12)).grid(row=3, column=0, padx=10, pady=5, sticky="e")
city_entry = tk.Entry(main_frame, textvariable=city_var, bg=entry_bg, font=('Arial', 12))
city_entry.grid(row=3, column=1, padx=10, pady=5, sticky="w")

tk.Label(main_frame, text="School Name:", bg=bg_color, fg=label_color, font=('Arial', 12)).grid(row=4, column=0, padx=10, pady=5, sticky="e")
school_name_entry = tk.Entry(main_frame, textvariable=school_name_var, bg=entry_bg, font=('Arial', 12))
school_name_entry.grid(row=4, column=1, padx=10, pady=5, sticky="w")

tk.Label(main_frame, text="Principal Name:", bg=bg_color, fg=label_color, font=('Arial', 12)).grid(row=5, column=0, padx=10, pady=5, sticky="e")
principal_name_entry = tk.Entry(main_frame, textvariable=principal_name_var, bg=entry_bg, font=('Arial', 12))
principal_name_entry.grid(row=5, column=1, padx=10, pady=5, sticky="w")

tk.Label(main_frame, text="Address:", bg=bg_color, fg=label_color, font=('Arial', 12)).grid(row=6, column=0, padx=10, pady=5, sticky="e")
address_entry = tk.Entry(main_frame, textvariable=address_var, bg=entry_bg, font=('Arial', 12))
address_entry.grid(row=6, column=1, padx=10, pady=5, sticky="w")

tk.Label(main_frame, text="Phone Number:", bg=bg_color, fg=label_color, font=('Arial', 12)).grid(row=7, column=0, padx=10, pady=5, sticky="e")
phone_number_entry = tk.Entry(main_frame, textvariable=phone_number_var, bg=entry_bg, font=('Arial', 12))
phone_number_entry.grid(row=7, column=1, padx=10, pady=5, sticky="w")

tk.Label(main_frame, text="Email:", bg=bg_color, fg=label_color, font=('Arial', 12)).grid(row=8, column=0, padx=10, pady=5, sticky="e")
email_entry = tk.Entry(main_frame, textvariable=email_var, bg=entry_bg, font=('Arial', 12))
email_entry.grid(row=8, column=1, padx=10, pady=5, sticky="w")

# Buttons for updating, deleting, and clearing
# Buttons for updating, deleting, and clearing
update_button = tk.Button(main_frame, text="Update/Insert School", command=update_school, bg="#1E90FF", fg="white", font=('Arial', 10))
update_button.grid(row=9, column=0, padx=10, pady=5)

delete_button = tk.Button(main_frame, text="Delete School", command=delete_school, bg="red", fg="white", font=('Arial', 10))
delete_button.grid(row=9, column=1, padx=10, pady=5)

clear_button = tk.Button(main_frame, text="Clear Fields", command=lambda: [sr_no_var.set(""), city_var.set(""), school_name_var.set(""), principal_name_var.set(""), address_var.set(""), phone_number_var.set(""), email_var.set(""), set_field_colors("")], bg="orange", fg="white", font=('Arial', 10))
clear_button.grid(row=9, column=2, padx=10, pady=5)

# Start Tkinter loop
root.mainloop()
